"""
web_scrappie -- a desktop GUI tool for scraping product images and metadata
from e-commerce websites. Reads category/URL pairs from spreadsheets (.ods,
.xlsx) or PDF files, visits each page with a real browser, pulls out every
product image and title it can find, and writes everything to a formatted
Excel workbook with one sheet per category.

Supports threaded image downloads and optional image embedding in Excel cells.

Author: Gopal Bagaswar
License: MIT
"""

import os
import sys
import subprocess
import threading
import ssl
import time
import re
import hashlib
import logging
import json
import csv
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse
from concurrent.futures import ThreadPoolExecutor, as_completed

# -- ssl fix for restricted networks ------------------------------------------

try:
    ssl._create_default_https_context = ssl._create_unverified_context
except AttributeError:
    pass
os.environ["PYTHONHTTPSVERIFY"] = "0"

# -- auto-install missing packages -------------------------------------------

DEPS = {
    "pandas": "pandas", "odfpy": "odfpy", "openpyxl": "openpyxl",
    "selenium": "selenium", "undetected_chromedriver": "undetected-chromedriver",
    "requests": "requests", "lxml": "lxml", "certifi": "certifi",
    "PIL": "Pillow", "pdfplumber": "pdfplumber", "customtkinter": "customtkinter",
}

def _bootstrap():
    missing = []
    for mod, pip_name in DEPS.items():
        try:
            __import__(mod)
        except ImportError:
            missing.append(pip_name)
    if missing:
        print(f"Installing: {', '.join(missing)}")
        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", "-q",
             "--trusted-host", "pypi.org",
             "--trusted-host", "files.pythonhosted.org"] + missing)

_bootstrap()

# -- imports (safe after bootstrap) -------------------------------------------

import pandas as pd
import requests
import warnings
import pdfplumber
import customtkinter as ctk
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage

warnings.filterwarnings("ignore", message="Unverified HTTPS request")
requests.packages.urllib3.disable_warnings()

log = logging.getLogger("web_scrappie")
log.setLevel(logging.INFO)

APP_VERSION = "1.2.0"

from tkinter import messagebox


# =============================================================================
#  Browser
# =============================================================================

def create_browser(headless=False):
    """Start a Chrome instance. Tries undetected-chromedriver first, then
    falls back to vanilla selenium."""

    import urllib.request
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    urllib.request.install_opener(
        urllib.request.build_opener(urllib.request.HTTPSHandler(context=ctx)))

    ua = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
          "AppleWebKit/537.36 (KHTML, like Gecko) "
          "Chrome/121.0.0.0 Safari/537.36")

    # -- attempt 1: undetected-chromedriver -----------------------------------
    uc_err = None
    try:
        import undetected_chromedriver as uc
        opts = uc.ChromeOptions()
        if headless:
            opts.add_argument("--headless=new")
            opts.add_argument("--disable-gpu")
            opts.add_argument(f"--user-agent={ua}")
        opts.add_argument("--no-sandbox")
        opts.add_argument("--disable-dev-shm-usage")
        opts.add_argument("--window-size=1920,1080")
        opts.add_argument("--disable-blink-features=AutomationControlled")
        opts.add_argument("--ignore-certificate-errors")
        opts.add_argument("--lang=en-US")
        driver = uc.Chrome(options=opts, version_main=None)
        if headless:
            driver.execute_cdp_cmd("Network.setUserAgentOverride",
                                   {"userAgent": ua, "platform": "Win32"})
            driver.execute_script(
                "Object.defineProperty(navigator,'webdriver',{get:()=>undefined});"
                "Object.defineProperty(navigator,'plugins',{get:()=>[1,2,3]});"
                "Object.defineProperty(navigator,'languages',{get:()=>['en-US','en']});"
                "window.chrome={runtime:{}};"
            )
        driver.set_page_load_timeout(60)
        driver.implicitly_wait(5)
        log.info("browser ready (uc, headless=%s)", headless)
        return driver
    except Exception as exc:
        uc_err = exc
        log.warning("undetected-chromedriver failed: %s", exc)

    # -- attempt 2: vanilla selenium ------------------------------------------
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    opts = Options()
    if headless:
        opts.add_argument("--headless=new")
        opts.add_argument("--disable-gpu")
        opts.add_argument(f"--user-agent={ua}")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1920,1080")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_argument("--ignore-certificate-errors")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    driver = webdriver.Chrome(options=opts)
    driver.execute_script(
        "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})")
    driver.set_page_load_timeout(60)
    driver.implicitly_wait(5)
    log.info("browser ready (selenium, headless=%s)", headless)
    return driver


# =============================================================================
#  Input file readers (.ods, .xlsx, .pdf)
# =============================================================================

def read_input_file(filepath):
    """Return {category: [url, ...]} from the given file."""
    ext = Path(filepath).suffix.lower()
    if ext == ".ods":
        return _read_spreadsheet(filepath, engine="odf")
    if ext in (".xlsx", ".xls"):
        return _read_spreadsheet(filepath, engine="openpyxl")
    if ext == ".pdf":
        return _read_pdf(filepath)
    raise ValueError(f"Unsupported format: {ext}  (use .ods, .xlsx, or .pdf)")


def _find_col(df, keywords):
    for col in df.columns:
        if any(kw in col for kw in keywords):
            return col
    return None


def _read_spreadsheet(filepath, engine):
    sheets = pd.read_excel(filepath, engine=engine, sheet_name=None)
    result = {}
    for sheet_name, df in sheets.items():
        log.info("  sheet '%s' -- %d rows, cols: %s",
                 sheet_name, len(df), list(df.columns))
        df.columns = [str(c).strip().lower() for c in df.columns]

        cat_col = _find_col(df, ["category", "cat", "type", "group", "class", "label"])
        url_col = _find_col(df, ["url", "link", "href", "web", "address", "site"])

        if cat_col is None and url_col is None and len(df.columns) == 2:
            cat_col, url_col = df.columns[0], df.columns[1]

        if url_col is None:
            for col in df.columns:
                vals = df[col].dropna().astype(str).head(10)
                if vals.str.contains(r"https?://", case=False).any():
                    url_col = col
                    break
        if url_col is None:
            continue

        for _, row in df.iterrows():
            cat = str(row.get(cat_col, sheet_name)).strip() if cat_col else sheet_name
            raw = str(row[url_col]).strip()
            if not raw or raw.lower() == "nan" or not raw.startswith("http"):
                continue
            result.setdefault(cat, []).append(raw)
    return result


def _read_pdf(filepath):
    result = {}
    url_re = re.compile(r"https?://[^\s\)\]\"'<>]+")
    with pdfplumber.open(filepath) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for u in url_re.findall(text):
                u = u.rstrip(".,;:")
                result.setdefault("PDF Links", []).append(u)

    # try PyMuPDF for annotation links
    try:
        import fitz
        doc = fitz.open(filepath)
        for page in doc:
            for link in page.get_links():
                uri = link.get("uri", "")
                if uri.startswith("http"):
                    result.setdefault("PDF Links", []).append(uri)
        doc.close()
    except Exception:
        pass

    # dedupe
    for cat in result:
        seen, unique = set(), []
        for u in result[cat]:
            if u not in seen:
                seen.add(u)
                unique.append(u)
        result[cat] = unique
    return result


# =============================================================================
#  Scraping
# =============================================================================

def scrape_page(driver, url, cfg, stop_flag):
    """Load a URL, scroll to trigger lazy-loading, then pull every product
    image it can find along with its title text."""

    driver.get(url)
    time.sleep(cfg["page_load_wait"])

    last_h = driver.execute_script("return document.body.scrollHeight")
    for _ in range(cfg["max_scroll"]):
        if stop_flag():
            break
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
        time.sleep(cfg["scroll_pause"])
        try:
            driver.execute_script("""
                for (const el of document.querySelectorAll('button, a')) {
                    const t = el.textContent.toLowerCase().trim();
                    if (['show more','load more','view more','next page',
                         'see more','view all'].some(k => t.includes(k))) {
                        el.click(); break;
                    }
                }""")
        except Exception:
            pass
        new_h = driver.execute_script("return document.body.scrollHeight")
        if new_h == last_h:
            break
        last_h = new_h

    raw = driver.execute_script("""
        const out = [], seen = new Set();

        // method 1: structured state (React / Next.js sites)
        const state = window.__PRELOADED_STATE__ || window.__NEXT_DATA__
                   || window.__INITIAL_STATE__ || null;
        if (state) {
            const items = (state.products)
                || (state.productList && state.productList.products)
                || (state.props && state.props.pageProps && state.props.pageProps.products)
                || [];
            for (const p of items) {
                const name = p.name || p.productName || p.title || '';
                let img = '';
                if (p.image && p.image.url) img = p.image.url;
                else if (p.imageUrl) img = p.imageUrl;
                else if (p.imageSrc) img = p.imageSrc;
                else if (p.images && p.images[0]) img = p.images[0].url || p.images[0];
                const link = p.url || p.productUrl || p.pdpUrl || '';
                if (name && img && !seen.has(name)) {
                    seen.add(name);
                    out.push({title: name, image_url: img, page_url: link});
                }
            }
            if (out.length) return out;
        }

        // method 2: product links wrapping images
        for (const a of document.querySelectorAll('a[href]')) {
            const img = a.querySelector('img');
            if (!img) continue;
            const src = img.src || img.dataset.src
                     || img.getAttribute('data-lazy-src') || '';
            if (!src || src.length < 10) continue;
            if (img.naturalWidth > 0 && img.naturalWidth < 40) continue;
            const title = img.alt || a.title || a.getAttribute('aria-label') || '';
            const key = src.split('?')[0];
            if (title && !seen.has(key)) {
                seen.add(key);
                out.push({title: title.substring(0, 200),
                          image_url: key, page_url: a.href || ''});
            }
        }
        if (out.length) return out;

        // method 3: standalone images with alt text
        for (const img of document.querySelectorAll('img[alt]')) {
            const src = (img.src || '').split('?')[0];
            if (!src || src.length < 10) continue;
            if (img.naturalWidth > 0 && img.naturalWidth < 40) continue;
            const title = img.alt || '';
            if (title && !seen.has(src)) {
                seen.add(src);
                const parent = img.closest('a');
                out.push({title: title.substring(0, 200),
                          image_url: src, page_url: parent ? parent.href : ''});
            }
        }
        return out;
    """)

    base = f"{urlparse(url).scheme}://{urlparse(url).netloc}"
    results = []
    for item in (raw or []):
        img = item.get("image_url", "")
        if img and not img.startswith("http"):
            img = ("https:" + img) if img.startswith("//") else base + img
        purl = item.get("page_url", "")
        if purl and not purl.startswith("http"):
            purl = base + purl
        results.append({
            "title": item.get("title", "").strip(),
            "image_url": img,
            "page_url": purl,
        })
    return results


def scrape_safe(driver, url, cfg, stop_flag, retries=2):
    for attempt in range(retries + 1):
        try:
            return scrape_page(driver, url, cfg, stop_flag)
        except Exception as exc:
            if attempt < retries:
                log.warning("  retry %d: %s", attempt + 1, exc)
                time.sleep(3)
            else:
                log.error("  gave up: %s", exc)
                return []


# =============================================================================
#  Image downloader (threaded)
# =============================================================================

def download_image(img_url, folder, prefix=""):
    os.makedirs(folder, exist_ok=True)
    h = hashlib.md5(img_url.encode()).hexdigest()[:8]
    safe = re.sub(r"[^\w\-]", "_", prefix)[:50]
    dest = os.path.join(folder, f"{safe}_{h}.jpg")
    if os.path.exists(dest):
        return dest
    try:
        r = requests.get(img_url, timeout=15, verify=False,
                         headers={"User-Agent": "Mozilla/5.0"})
        if r.status_code == 200 and len(r.content) > 500:
            with open(dest, "wb") as f:
                f.write(r.content)
            return dest
    except Exception:
        pass
    return ""


# =============================================================================
#  Excel output
# =============================================================================

def save_to_excel(all_data, output_path, download_images=False):
    wb = Workbook()
    wb.remove(wb.active)

    H_FONT  = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    H_FILL  = PatternFill("solid", fgColor="2B547E")
    H_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    C_FONT  = Font(name="Arial", size=10)
    C_ALIGN = Alignment(vertical="top", wrap_text=True)
    L_FONT  = Font(name="Arial", size=10, color="0563C1", underline="single")
    BRD     = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"))
    ALT = PatternFill("solid", fgColor="F2F7FB")

    if download_images:
        headers = ["#", "Image", "Title", "Image URL", "Page URL"]
        widths  = [6, 15, 50, 55, 55]
        CW, CH, RH = 100, 95, 75
        CI, CT, CU, CP = 2, 3, 4, 5
    else:
        headers = ["#", "Title", "Image URL", "Page URL"]
        widths  = [6, 55, 60, 60]
        CT, CU, CP = 2, 3, 4

    total = 0
    for cat, items in sorted(all_data.items()):
        if not items:
            continue
        sname = re.sub(r"[\\/*?\[\]:]", "", cat)[:31]
        ws = wb.create_sheet(title=sname)

        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font, c.fill, c.alignment, c.border = H_FONT, H_FILL, H_ALIGN, BRD
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        for ri, item in enumerate(items, 2):
            ws.cell(row=ri, column=1, value=ri - 1).font = C_FONT
            ws.cell(row=ri, column=1).alignment = Alignment(
                horizontal="center", vertical="center")

            ws.cell(row=ri, column=CT, value=item.get("title", "")).font = C_FONT
            ws.cell(row=ri, column=CT).alignment = Alignment(
                vertical="center", wrap_text=True)

            iurl = item.get("image_url", "")
            c = ws.cell(row=ri, column=CU, value=iurl)
            c.font, c.alignment = L_FONT, C_ALIGN
            if iurl:
                c.hyperlink = iurl

            purl = item.get("page_url", "")
            c = ws.cell(row=ri, column=CP, value=purl)
            c.font, c.alignment = L_FONT, C_ALIGN
            if purl:
                c.hyperlink = purl

            if download_images:
                lp = item.get("local_path", "")
                if lp and os.path.exists(lp):
                    try:
                        img = XlImage(lp)
                        ow, oh = img.width or 1, img.height or 1
                        pad = 4
                        ratio = min((CW - pad * 2) / ow, (CH - pad * 2) / oh)
                        img.width  = int(ow * ratio)
                        img.height = int(oh * ratio)
                        ws.add_image(img, f"{get_column_letter(CI)}{ri}")
                        ws.row_dimensions[ri].height = RH
                    except Exception:
                        ws.cell(row=ri, column=CI, value="(err)").font = C_FONT

            if ri % 2 == 0:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=ri, column=col).fill = ALT
            for col in range(1, len(headers) + 1):
                ws.cell(row=ri, column=col).border = BRD

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(items) + 1}"
        total += len(items)

    # summary sheet
    ws_s = wb.create_sheet(title="Summary", index=0)
    ws_s["A1"] = "web_scrappie -- run summary"
    ws_s["A1"].font = Font(name="Arial", bold=True, size=14, color="2B547E")
    ws_s["A3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_s["A3"].font = Font(name="Arial", size=10, color="666666")
    ws_s["A5"], ws_s["B5"] = "Category", "Items Found"
    ws_s["A5"].font = ws_s["B5"].font = H_FONT
    ws_s["A5"].fill = ws_s["B5"].fill = H_FILL
    ws_s.column_dimensions["A"].width = 35
    ws_s.column_dimensions["B"].width = 18
    for i, (cat, items) in enumerate(sorted(all_data.items()), 6):
        ws_s.cell(row=i, column=1, value=cat).font = C_FONT
        ws_s.cell(row=i, column=2, value=len(items)).font = C_FONT
    r = 6 + len(all_data)
    ws_s.cell(row=r, column=1, value="TOTAL").font = Font(
        name="Arial", bold=True, size=10)
    ws_s.cell(row=r, column=2, value=total).font = Font(
        name="Arial", bold=True, size=10)

    wb.save(output_path)
    return total


def save_to_csv(all_data, output_path):
    """Save scraped data as a flat CSV with columns:
    Category, Title, Image URL, Page URL."""
    csv_path = os.path.splitext(output_path)[0] + ".csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Category", "Title", "Image URL", "Page URL"])
        for cat, items in sorted(all_data.items()):
            for item in items:
                writer.writerow([
                    cat,
                    item.get("title", ""),
                    item.get("image_url", ""),
                    item.get("page_url", ""),
                ])
    return csv_path


def save_to_json(all_data, output_path):
    """Save scraped data as structured JSON with categories as keys."""
    json_path = os.path.splitext(output_path)[0] + ".json"
    structured = {}
    for cat, items in sorted(all_data.items()):
        structured[cat] = [
            {
                "title": item.get("title", ""),
                "image_url": item.get("image_url", ""),
                "page_url": item.get("page_url", ""),
            }
            for item in items
        ]
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(structured, f, indent=2, ensure_ascii=False)
    return json_path


# =============================================================================
#  Scraper engine (background thread)
# =============================================================================

def run_engine(cfg, log_cb, done_cb, stop_flag):
    def emit(msg):
        log.info(msg)
        log_cb(msg)

    driver = None
    try:
        emit("reading input file...")
        data = read_input_file(cfg["input_file"])
        if not data:
            emit("ERROR: no valid URLs found in file.")
            done_cb(False)
            return

        n_urls = sum(len(v) for v in data.values())
        emit(f"found {n_urls} URLs across {len(data)} categories")
        for cat, urls in data.items():
            emit(f"  {cat}: {len(urls)} links")

        emit("\nstarting browser...")
        if cfg["headless"]:
            emit("(headless mode -- some sites may block this)")
        driver = create_browser(headless=cfg["headless"])

        # warm up: visit the first domain to set cookies
        first_url = next(iter(next(iter(data.values()))))
        domain = f"{urlparse(first_url).scheme}://{urlparse(first_url).netloc}"
        emit(f"warming up on {domain}...")
        try:
            driver.get(domain)
            time.sleep(3)
        except Exception:
            pass

        all_results = {}
        url_idx = 0

        for category, urls in data.items():
            if stop_flag():
                emit("\nstopped by user.")
                break
            emit(f"\n{'=' * 50}")
            emit(f"category: {category} ({len(urls)} URLs)")
            cat_items = []

            for url in urls:
                if stop_flag():
                    emit("\nstopped by user.")
                    break
                url_idx += 1
                emit(f"  [{url_idx}/{n_urls}] {url[:70]}...")

                items = scrape_safe(driver, url, cfg, stop_flag)
                emit(f"    found {len(items)} items")

                if cfg["download_images"] and items:
                    emit(f"    downloading images ({cfg['threads']} threads)...")
                    folder = os.path.join(cfg["image_folder"],
                                          re.sub(r"[^\w\-]", "_", category))
                    os.makedirs(folder, exist_ok=True)

                    def _dl(p, f=folder):
                        p["local_path"] = download_image(
                            p["image_url"], f, p["title"][:30])
                        return bool(p["local_path"])

                    ok = 0
                    with ThreadPoolExecutor(max_workers=cfg["threads"]) as pool:
                        futs = {pool.submit(_dl, p): p for p in items}
                        for fut in as_completed(futs):
                            try:
                                if fut.result():
                                    ok += 1
                            except Exception:
                                pass
                    emit(f"    downloaded {ok}/{len(items)}")

                cat_items.extend(items)
                time.sleep(1)

            # deduplicate by image url
            seen, unique = set(), []
            for p in cat_items:
                key = p.get("image_url", "")
                if key and key not in seen:
                    seen.add(key)
                    unique.append(p)
            all_results[category] = unique
            emit(f"  total for '{category}': {len(unique)} unique items")

        if driver:
            driver.quit()
            emit("\nbrowser closed.")

        if all_results:
            out = cfg["output_file"]
            out_dir = os.path.dirname(out)
            if out_dir:
                os.makedirs(out_dir, exist_ok=True)
            emit(f"saving: {out}")
            total = save_to_excel(all_results, out, cfg["download_images"])
            csv_path = save_to_csv(all_results, out)
            json_path = save_to_json(all_results, out)
            emit(f"\ndone. {total} items across {len(all_results)} categories.")
            emit(f"file: {out}")
            emit(f"csv:  {csv_path}")
            emit(f"json: {json_path}")
            done_cb(True)
        else:
            emit("\nnothing scraped.")
            done_cb(False)

    except Exception as exc:
        emit(f"\nERROR: {exc}")
        if driver:
            try:
                driver.quit()
            except Exception:
                pass
        done_cb(False)


# =============================================================================
#  GUI (customtkinter)
# =============================================================================

# -- colour palette -----------------------------------------------------------
_ACCENT     = "#3B82F6"   # vibrant blue
_ACCENT_H   = "#2563EB"   # hover
_DANGER     = "#EF4444"   # red
_DANGER_H   = "#DC2626"
_SUCCESS    = "#22C55E"
_SURFACE    = "#1E293B"   # card bg (dark mode)
_SURFACE_L  = "#F1F5F9"   # card bg (light mode)
_MUTED      = "#94A3B8"
_BG_DARK    = "#0F172A"


class App(ctk.CTk):

    def __init__(self):
        super().__init__()
        self.title(f"web_scrappie v{APP_VERSION}")
        self.geometry("860x760")
        self.minsize(720, 620)
        self._stop = False
        self._url_count = 0
        self._url_done = 0

        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")

        self.configure(fg_color=_BG_DARK)
        self._build()

    # -- layout ---------------------------------------------------------------

    def _build(self):

        # ---- header ---------------------------------------------------------
        header = ctk.CTkFrame(self, fg_color="transparent")
        header.pack(fill="x", padx=28, pady=(22, 0))

        title_col = ctk.CTkFrame(header, fg_color="transparent")
        title_col.pack(side="left")

        ctk.CTkLabel(
            title_col, text="web_scrappie",
            font=ctk.CTkFont(size=28, weight="bold"),
            text_color="#F8FAFC",
        ).pack(anchor="w")

        ctk.CTkLabel(
            title_col,
            text="Scrape product images & metadata from any e-commerce site",
            font=ctk.CTkFont(size=13), text_color=_MUTED,
        ).pack(anchor="w", pady=(2, 0))

        # theme toggle (sun/moon)
        self._dark_mode = True
        self.theme_btn = ctk.CTkButton(
            header, text="\u263D  Dark", width=80, height=28,
            corner_radius=8, fg_color="#334155", hover_color="#475569",
            font=ctk.CTkFont(size=12), command=self._toggle_theme,
        )
        self.theme_btn.pack(side="right", padx=(8, 0), pady=(4, 0))

        # version badge
        badge = ctk.CTkFrame(header, fg_color=_ACCENT, corner_radius=12,
                              width=60, height=26)
        badge.pack(side="right", pady=(4, 0))
        badge.pack_propagate(False)
        ctk.CTkLabel(badge, text=f"v{APP_VERSION}",
                     font=ctk.CTkFont(size=11, weight="bold"),
                     text_color="#FFFFFF").place(relx=0.5, rely=0.5,
                                                 anchor="center")

        # thin separator
        ctk.CTkFrame(self, height=1, fg_color="#334155").pack(
            fill="x", padx=28, pady=(14, 0))

        # scrollable body
        body = ctk.CTkScrollableFrame(self, fg_color="transparent",
                                       label_text="")
        body.pack(fill="both", expand=True, padx=20, pady=(8, 0))

        # ---- input card -----------------------------------------------------
        self._card_input = self._card(body, "Input File")
        inner = self._card_input

        ctk.CTkLabel(
            inner, text="Supported: .ods  .xlsx  .xls  .pdf",
            font=ctk.CTkFont(size=11), text_color=_MUTED,
        ).pack(anchor="w", padx=14, pady=(0, 6))

        row = ctk.CTkFrame(inner, fg_color="transparent")
        row.pack(fill="x", padx=14, pady=(0, 12))
        self.file_var = ctk.StringVar()
        self._file_entry = ctk.CTkEntry(
            row, textvariable=self.file_var,
            placeholder_text="Choose a spreadsheet or PDF...",
            height=38, corner_radius=8,
        )
        self._file_entry.pack(side="left", fill="x", expand=True)
        ctk.CTkButton(
            row, text="Browse", width=90, height=38, corner_radius=8,
            fg_color=_ACCENT, hover_color=_ACCENT_H,
            command=self._browse_input,
        ).pack(side="left", padx=(8, 0))

        # preview URL count label
        self.preview_label = ctk.CTkLabel(
            inner, text="",
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color=_SUCCESS,
        )
        self.preview_label.pack(anchor="w", padx=14, pady=(0, 8))

        # drag & drop support (requires tkinterdnd2)
        try:
            self._file_entry.drop_target_register("DND_Files")
            self._file_entry.dnd_bind("<<Drop>>", self._on_drop)
        except Exception:
            pass  # tkinterdnd2 not available, skip gracefully

        # ---- settings card --------------------------------------------------
        settings = self._card(body, "Scraper Settings")

        grid = ctk.CTkFrame(settings, fg_color="transparent")
        grid.pack(fill="x", padx=14, pady=(0, 4))
        grid.columnconfigure((0, 1, 2, 3), weight=1)

        self._add_field(grid, "Max Scrolls", "15", 0, 0)
        self._add_field(grid, "Scroll Pause (s)", "2.0", 0, 1)
        self._add_field(grid, "Page Wait (s)", "8", 0, 2)
        self._add_field(grid, "DL Threads", "8", 0, 3)

        toggles = ctk.CTkFrame(settings, fg_color="transparent")
        toggles.pack(fill="x", padx=14, pady=(4, 12))

        self.headless_var = ctk.BooleanVar(value=False)
        ctk.CTkSwitch(
            toggles, text="Headless mode  (faster but may get blocked)",
            variable=self.headless_var, onvalue=True, offvalue=False,
            progress_color=_ACCENT,
        ).pack(side="left", padx=(0, 24))

        self.dl_var = ctk.BooleanVar(value=False)
        ctk.CTkSwitch(
            toggles, text="Download images & embed in Excel",
            variable=self.dl_var, onvalue=True, offvalue=False,
            progress_color=_ACCENT,
        ).pack(side="left")

        # ---- output card ----------------------------------------------------
        out_card = self._card(body, "Output")
        row_o = ctk.CTkFrame(out_card, fg_color="transparent")
        row_o.pack(fill="x", padx=14, pady=(0, 12))

        self.out_var = ctk.StringVar(value=os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            f"scrape_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"))
        ctk.CTkEntry(
            row_o, textvariable=self.out_var, height=38, corner_radius=8,
        ).pack(side="left", fill="x", expand=True)
        ctk.CTkButton(
            row_o, text="Browse", width=90, height=38, corner_radius=8,
            fg_color=_ACCENT, hover_color=_ACCENT_H,
            command=self._browse_output,
        ).pack(side="left", padx=(8, 0))

        # ---- action bar -----------------------------------------------------
        action = ctk.CTkFrame(self, fg_color="transparent")
        action.pack(fill="x", padx=28, pady=(10, 0))

        self.start_btn = ctk.CTkButton(
            action, text="Start Scraping", width=200, height=44,
            corner_radius=10, fg_color=_ACCENT, hover_color=_ACCENT_H,
            font=ctk.CTkFont(size=15, weight="bold"),
            command=self._start,
        )
        self.start_btn.pack(side="left")

        self.stop_btn = ctk.CTkButton(
            action, text="Stop", width=110, height=44,
            corner_radius=10,
            fg_color=_DANGER, hover_color=_DANGER_H,
            font=ctk.CTkFont(size=15, weight="bold"),
            state="disabled", command=self._request_stop,
        )
        self.stop_btn.pack(side="left", padx=(10, 0))

        self.open_btn = ctk.CTkButton(
            action, text="Open File", width=110, height=44,
            corner_radius=10,
            fg_color=_SUCCESS, hover_color="#16A34A",
            font=ctk.CTkFont(size=15, weight="bold"),
            state="disabled", command=self._open_output,
        )
        self.open_btn.pack(side="left", padx=(10, 0))

        # progress bar
        self.progress = ctk.CTkProgressBar(
            action, width=200, height=14, corner_radius=7,
            progress_color=_ACCENT,
        )
        self.progress.pack(side="right", padx=(10, 0))
        self.progress.set(0)

        self.progress_label = ctk.CTkLabel(
            action, text="0 / 0 URLs",
            font=ctk.CTkFont(size=12), text_color=_MUTED,
        )
        self.progress_label.pack(side="right")

        # ---- log card -------------------------------------------------------
        log_frame = ctk.CTkFrame(self, fg_color=_SURFACE, corner_radius=12)
        log_frame.pack(fill="both", expand=True, padx=20, pady=(10, 6))

        log_header = ctk.CTkFrame(log_frame, fg_color="transparent")
        log_header.pack(fill="x", padx=12, pady=(10, 4))
        ctk.CTkLabel(
            log_header, text="Live Log",
            font=ctk.CTkFont(size=12, weight="bold"),
        ).pack(side="left")
        ctk.CTkButton(
            log_header, text="Clear", width=55, height=24,
            corner_radius=6, fg_color="#334155", hover_color="#475569",
            font=ctk.CTkFont(size=11), command=self._clear_log,
        ).pack(side="left", padx=(8, 0))

        self.log_box = ctk.CTkTextbox(
            log_frame,
            font=ctk.CTkFont(family="Menlo", size=12),
            fg_color="#0F172A", text_color="#CBD5E1",
            state="disabled", wrap="word", corner_radius=8,
        )
        self.log_box.pack(fill="both", expand=True, padx=8, pady=(0, 8))

        # ---- status bar -----------------------------------------------------
        status_bar = ctk.CTkFrame(self, fg_color="#1E293B", height=30,
                                   corner_radius=0)
        status_bar.pack(fill="x", side="bottom")
        status_bar.pack_propagate(False)

        self.status_var = ctk.StringVar(value="Ready")
        ctk.CTkLabel(
            status_bar, textvariable=self.status_var,
            font=ctk.CTkFont(size=11), text_color=_MUTED,
        ).pack(side="left", padx=16)

        ctk.CTkLabel(
            status_bar, text="by Gopal Bagaswar",
            font=ctk.CTkFont(size=11), text_color="#475569",
        ).pack(side="right", padx=16)

    # -- helpers --------------------------------------------------------------

    def _card(self, parent, title):
        """Create a styled card frame with a title."""
        frame = ctk.CTkFrame(parent, fg_color=_SURFACE, corner_radius=12)
        frame.pack(fill="x", pady=(8, 0))
        ctk.CTkLabel(
            frame, text=title,
            font=ctk.CTkFont(size=13, weight="bold"),
        ).pack(anchor="w", padx=14, pady=(12, 6))
        return frame

    def _add_field(self, parent, label, default, row, col):
        """Add a labelled entry field in a grid layout."""
        cell = ctk.CTkFrame(parent, fg_color="transparent")
        cell.grid(row=row, column=col, padx=(0, 12), pady=4, sticky="ew")
        ctk.CTkLabel(
            cell, text=label,
            font=ctk.CTkFont(size=11), text_color=_MUTED,
        ).pack(anchor="w")
        var = ctk.StringVar(value=default)
        ctk.CTkEntry(cell, textvariable=var, width=80, height=32,
                     corner_radius=6).pack(anchor="w", pady=(2, 0))
        key = label.replace(" ", "_").replace("(", "").replace(")", "").lower()
        setattr(self, f"_field_{key}", var)
        return var

    def _browse_input(self):
        path = ctk.filedialog.askopenfilename(
            title="Select Input File",
            filetypes=[("Spreadsheets", "*.ods *.xlsx *.xls"),
                       ("PDF", "*.pdf"),
                       ("All", "*.*")])
        if path:
            self.file_var.set(path)
            self._preview_file(path)

    def _browse_output(self):
        path = ctk.filedialog.asksaveasfilename(
            title="Save As", defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")])
        if path:
            self.out_var.set(path)

    def _preview_file(self, path):
        """Parse the input file and show a summary of URLs found."""
        try:
            data = read_input_file(path)
            n_urls = sum(len(v) for v in data.values())
            n_cats = len(data)
            self.preview_label.configure(
                text=f"Found {n_urls} URLs across {n_cats} "
                     f"categor{'y' if n_cats == 1 else 'ies'}")
        except Exception as exc:
            self.preview_label.configure(
                text=f"Could not parse file: {exc}",
                text_color=_DANGER)

    def _toggle_theme(self):
        """Switch between dark and light appearance mode."""
        self._dark_mode = not self._dark_mode
        if self._dark_mode:
            ctk.set_appearance_mode("dark")
            self.theme_btn.configure(text="\u263D  Dark")
        else:
            ctk.set_appearance_mode("light")
            self.theme_btn.configure(text="\u2600  Light")

    def _on_drop(self, event):
        """Handle drag-and-drop file onto the input entry."""
        path = event.data.strip().strip("{}")
        if os.path.isfile(path):
            self.file_var.set(path)
            self._preview_file(path)

    def _open_output(self):
        """Open the output Excel file with the system default application."""
        out = self.out_var.get().strip()
        if out and os.path.exists(out):
            try:
                subprocess.run(["open", out])
            except Exception as exc:
                messagebox.showerror("Error", f"Could not open file:\n{exc}")

    def _clear_log(self):
        """Clear the log textbox."""
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")

    def _log(self, msg):
        def _do():
            self.log_box.configure(state="normal")
            self.log_box.insert("end", msg + "\n")
            self.log_box.see("end")
            self.log_box.configure(state="disabled")
        self.after(0, _do)

    def _update_progress(self, done, total):
        def _do():
            frac = done / total if total else 0
            self.progress.set(frac)
            self.progress_label.configure(text=f"{done} / {total} URLs")
        self.after(0, _do)

    # -- actions --------------------------------------------------------------

    def _start(self):
        fp = self.file_var.get().strip()
        if not fp or not os.path.exists(fp):
            messagebox.showerror("Error", "Select a valid input file first.")
            return
        out = self.out_var.get().strip()
        if not out:
            messagebox.showerror("Error", "Set an output file path.")
            return

        def _get(name, fallback):
            attr = f"_field_{name}"
            v = getattr(self, attr, None)
            return v.get() if v else fallback

        cfg = {
            "input_file":     fp,
            "output_file":    out,
            "max_scroll":     int(_get("max_scrolls", "15")),
            "scroll_pause":   float(_get("scroll_pause_s", "2.0")),
            "page_load_wait": int(_get("page_wait_s", "8")),
            "threads":        int(_get("dl_threads", "8")),
            "headless":       self.headless_var.get(),
            "download_images": self.dl_var.get(),
            "image_folder":   os.path.join(os.path.dirname(out), "scraped_images"),
        }

        self._stop = False
        self._url_done = 0
        self.start_btn.configure(state="disabled")
        self.stop_btn.configure(state="normal")
        self.open_btn.configure(state="disabled")
        self.status_var.set("Running...")
        self.progress.set(0)
        self.progress_label.configure(text="starting...")

        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")

        # wrap log callback to detect progress
        _orig_log = self._log
        def _progress_log(msg):
            _orig_log(msg)
            if msg.strip().startswith("["):
                try:
                    inner = msg.strip().split("]")[0].lstrip("[").strip()
                    done, total = inner.split("/")
                    self._update_progress(int(done), int(total))
                except Exception:
                    pass

        def on_done(ok):
            self.after(0, lambda: self.start_btn.configure(state="normal"))
            self.after(0, lambda: self.stop_btn.configure(state="disabled"))
            if ok:
                self.after(0, lambda: self.progress.set(1.0))
                self.after(0, lambda: self.open_btn.configure(state="normal"))
                self.after(0, lambda: self.status_var.set(
                    f"Done -- {cfg['output_file']}"))
                self.after(0, lambda: messagebox.showinfo(
                    "Done", f"Scraping complete!\n\n{cfg['output_file']}"))
            else:
                self.after(0, lambda: self.status_var.set(
                    "Finished with errors"))

        t = threading.Thread(
            target=run_engine,
            args=(cfg, _progress_log, on_done, lambda: self._stop),
            daemon=True)
        t.start()

    def _request_stop(self):
        self._stop = True
        self._log("\nStop requested -- finishing current URL...")
        self.status_var.set("Stopping...")
        self.stop_btn.configure(state="disabled")


# =============================================================================
#  main
# =============================================================================

if __name__ == "__main__":
    app = App()
    app.mainloop()
